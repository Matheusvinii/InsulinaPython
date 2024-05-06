import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Funções para manipular a planilha Excel
def inicializar_planilha():
    wb = Workbook()
    ws = wb.active
    ws.append(["Data e Hora", "Entrada/Saída", "Atendente", "Quantidade", "Tipo de Insulina"])  # Adicionar coluna "Tipo de Insulina"
    return wb

def carregar_planilha():
    estoque_file = "estoque_insulina.xlsx"
    if not os.path.exists(estoque_file):
        wb = inicializar_planilha()
        wb.save(estoque_file)
    else:
        wb = load_workbook(estoque_file)
    return wb

# Função para atualizar o estoque e registrar a transação
def atualizar_estoque_e_registrar(quantidade, tipo, atendente, tipo_insulina):
    wb = carregar_planilha()
    ws = wb.active

    if tipo == "Entrada":
        if tipo_insulina == "NPH":
            st.session_state.estoque_nph += quantidade
        else:
            st.session_state.estoque_regular += quantidade
        st.session_state.estoque_total += quantidade  # Atualizar estoque total
    else:
        if tipo_insulina == "NPH":
            st.session_state.estoque_nph -= quantidade
        else:
            st.session_state.estoque_regular -= quantidade
        st.session_state.estoque_total -= quantidade  # Atualizar estoque total

    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws.append([agora, tipo, atendente, quantidade, tipo_insulina])
    wb.save("estoque_insulina.xlsx")
    salvar_estoques()

def carregar_estoques():
    estoque_file = "estoque_atual.txt"
    try:
        with open(estoque_file, "r") as f:
            estoque_nph, estoque_regular = map(int, f.read().split(","))
            estoque_total = estoque_nph + estoque_regular
    except FileNotFoundError:
        estoque_nph = 0
        estoque_regular = 0
        estoque_total = 0
    return estoque_nph, estoque_regular, estoque_total

def salvar_estoques():
    with open("estoque_atual.txt", "w") as f:
        f.write(f"{st.session_state.estoque_nph},{st.session_state.estoque_regular}")

# Inicializar o session state
if 'estoque_nph' not in st.session_state:
    st.session_state.estoque_nph, st.session_state.estoque_regular, st.session_state.estoque_total = carregar_estoques()

# Interface do usuário
st.title("Gerenciamento de Estoque de Insulina")

# Inserir o nome do atendente
atendente = st.text_input("Nome do Atendente:", "")

# Exibir os estoques
st.write(f"Estoque total: {st.session_state.estoque_total}")
st.write(f"Estoque NPH: {st.session_state.estoque_nph}")
st.write(f"Estoque Regular: {st.session_state.estoque_regular}")

# Opções de entrada e saída
opcao = st.radio("Selecione a operação:", ("Entrada", "Saída"))

# Inserir a quantidade
quantidade = st.number_input("Quantidade:", min_value=1, step=1)

# Selecionar o tipo de insulina
tipo_insulina = st.selectbox("Tipo de Insulina:", ("NPH", "Regular"))

# Botão para confirmar a operação
if st.button("Confirmar"):
    if not atendente:
        st.error("Por favor, insira o nome do atendente.")
    else:
        atualizar_estoque_e_registrar(quantidade, opcao, atendente, tipo_insulina)
        st.success("Estoque atualizado com sucesso!")
        st.experimental_rerun()

# Botão para exportar a planilha
st.download_button(
    label="Exportar Histórico",
    data=open("estoque_insulina.xlsx", "rb").read(),
    file_name="estoque_insulina.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)