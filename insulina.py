import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Funções para manipular a planilha Excel
def inicializar_planilha():
    wb = Workbook()
    ws = wb.active
    ws.append(["Data e Hora", "Entrada/Saída", "Atendente", "Quantidade"])  # Adicionar coluna "Quantidade"
    return wb


def carregar_planilha():
    estoque_file = "estoque_insulina.xlsx"
    if not os.path.exists(estoque_file):
        wb = inicializar_planilha()
        wb.save(estoque_file)
    else:
        wb = load_workbook(estoque_file)
    return wb

# Função para atualizar o estoque e registrar a transação
def atualizar_estoque_e_registrar(quantidade, tipo, atendente):
    wb = carregar_planilha()
    ws = wb.active

    if tipo == "Entrada":
        st.session_state.estoque += quantidade
    else:
        st.session_state.estoque -= quantidade

    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws.append([agora, tipo, atendente, quantidade])
    wb.save("estoque_insulina.xlsx")
    salvar_estoque(st.session_state.estoque)
    
def carregar_estoque():
    estoque_file = "estoque_atual.txt"
    try:
        with open(estoque_file, "r") as f:
            estoque = int(f.read())
    except FileNotFoundError:
        estoque = 0  # Valor inicial do estoque
    return estoque

# Inicializar o session state
if 'estoque' not in st.session_state:
    st.session_state.estoque = carregar_estoque()
    
def salvar_estoque(estoque):
    with open("estoque_atual.txt", "w") as f:
        f.write(str(estoque))

# Interface do usuário
st.title("Gerenciamento de Estoque de Insulina")

# Inserir o nome do atendente
atendente = st.text_input("Nome do Atendente:", "")

# Exibir o estoque atual
st.write(f"Estoque atual: {st.session_state.estoque}")

# Opções de entrada e saída
opcao = st.radio("Selecione a operação:", ("Entrada", "Saída"))

# Inserir a quantidade
quantidade = st.number_input("Quantidade:", min_value=1, step=1)

# Botão para confirmar a operação
if st.button("Confirmar"):
    if not atendente:
        st.error("Por favor, insira o nome do atendente.")
    else:
        atualizar_estoque_e_registrar(quantidade, opcao, atendente)
        st.success("Estoque atualizado com sucesso!")
        st.experimental_rerun()

# Botão para exportar a planilha
st.download_button(
    label="Exportar Histórico",
    data=open("estoque_insulina.xlsx", "rb").read(),
    file_name="estoque_insulina.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)